"""O motor de importação da planilha oficial de lavagens.

Vive aqui — e não em `scripts/` — porque tem dois chamadores: o endpoint de
upload do painel (`POST /importacoes`, via `processar_upload`) e o CLI
`scripts/importar_base.py`, que virou um invólucro fino. Não tem `index.py` na
pasta, então o auto-registro de rotas do `app.py` não o enxerga (mesmo
precedente de `use_cases/lavagem/contrato.py`).

O arquivo oficial é multi-aba — uma aba por ano ("Lavagens 2015"…). O motor
não depende dos nomes: processa toda aba cujo cabeçalho tenha as colunas
obrigatórias, e é assim que a aba do próximo ano entra sem mudança de código.
Os valores passam pela padronização de `utils/padronizar_valores.py` (data em
três formatos, preço com vírgula, 22 grafias de categoria), senão a carga
aborta ou o dashboard fatia errado.

Dois modos:

- **incremental** (padrão): `ON CONFLICT DO NOTHING` sobre `id_externo` — o
  arquivo inteiro volta, só o que é novo insere. Commit por lote.
- **substituir**: apaga as lavagens `origem='importacao'` (DELETE seletivo —
  os check-ins da plataforma ficam) e recarrega tudo NUMA TRANSAÇÃO SÓ: falha
  em qualquer ponto devolve o banco inteiro ao estado anterior. É o modo
  obrigatório da primeira carga da base nova, cujos ids coincidem com os da
  carga antiga.
"""
from dataclasses import dataclass, field
from datetime import date, datetime, time, timezone
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import openpyxl

from database.database import SessionLocal
from entities.lavagem import minutos_entre
from models.importacao_model import ImportacaoModel
from repositories.cliente_repository import ClienteRepository
from repositories.funcionario_repository import FuncionarioRepository
from repositories.lavagem_repository import LavagemRepository
from repositories.veiculo_repository import normalizar_placa
from models.cliente_model import ClienteModel
from models.veiculo_model import VeiculoModel
from models.lavagem_model import LavagemModel
from sqlalchemy import func
from utils.normalizar_texto import normalizar_cpf, normalizar_telefone, normalizar_texto
from utils.padronizar_valores import (
    METODOS_PAGAMENTO_CANONICOS,
    TIPOS_CARRO_CANONICOS,
    padronizar_categoria,
    padronizar_data,
    padronizar_preco,
)

# Sem estas duas não há linha: o id é a chave da idempotência e a data é o
# eixo de todo o dashboard. Aba sem elas no cabeçalho não é aba de lavagens.
COLUNAS_OBRIGATORIAS = {"id_lavagem", "data"}

# Colunas copiadas 1:1 para a coluna de mesmo nome na tabela.
COLUNAS_DIRETAS = (
    "tempo_espera_antes_lavagem_min",
    "tempo_lavagem_total_min",
    "tempo_pos_lavagem_ate_retirada_min",
    "tempo_ate_pagamento_min",
    "t_teto_vidros_min",
    "t_capo_parabrisa_min",
    "t_laterais_portas_min",
    "t_traseira_min",
    "t_rodas_pneus_min",
    "t_interior_min",
    "cnpj_receita",
    "nps_cliente",
    "nota_google",
    "shampoo_ml",
    "cera_ml",
    "pretinho_ml",
    "aromatizante_ml",
    "observacoes",
)

# Colunas com tratamento próprio (padronização, FK ou nome diferente no banco).
COLUNAS_ESPECIAIS = {
    "id_lavagem": "id_externo",
    "data": "data",
    "tipo_carro": "tipo_carro",
    "preco_reais": "preco_reais",
    "metodo_pagamento": "metodo_pagamento",
    "cpf": "cpf",
    "cliente": "cliente",
    "telefone": "telefone",
    "placa": "placa",
    "modelo_carro": "modelo",
    "funcionario_lavagem": "funcionario_lavagem",
    "atendente_pagamento": "atendente_pagamento",
    "servico": "servico",
    "chegou_em": "chegou_em",
    "iniciou_lavagem_em": "iniciou_lavagem_em",
    "terminou_lavagem_em": "terminou_lavagem_em",
    "saiu_em": "saiu_em",
    "pago_em": "pago_em",
}

# Apelidos: como as colunas que hoje faltam provavelmente virão escritas.
APELIDOS = {
    "telefone_cliente": "telefone",
    "celular": "telefone",
    "whatsapp": "telefone",
    "placa_veiculo": "placa",
    "modelo": "modelo_carro",
    "horario_chegada": "chegou_em",
    "hora_chegada": "chegou_em",
    "chegada": "chegou_em",
    "horario_entrada": "chegou_em",
    "horario_inicio_lavagem": "iniciou_lavagem_em",
    "horario_fim_lavagem": "terminou_lavagem_em",
    "horario_saida": "saiu_em",
    "hora_saida": "saiu_em",
    "saida": "saiu_em",
    "horario_pagamento": "pago_em",
    "nome_cliente": "cliente",
    "lavador": "funcionario_lavagem",
    "atendente": "atendente_pagamento",
    "valor": "preco_reais",
    "valor_reais": "preco_reais",
    "preco": "preco_reais",
    "nps": "nps_cliente",
}

# Derivadas de `data` — não viram coluna (o dashboard usa date_trunc), mas
# `ano` e `mes` são LIDAS antes do descarte: servem de gabarito para
# desambiguar a data em texto com barra (dd/mm vs mm/dd).
COLUNAS_IGNORADAS = ("ano", "mes", "mês", "dia_semana")

# O que a imputação antiga escrevia no lugar do nome que faltava.
NOME_PLACEHOLDER = "nao informado"

LOTE_PADRAO = 5000

# Teto de exemplos de grafia desconhecida por aba no resumo — o dono precisa
# ver que existe, não de uma lista com centenas de repetições.
MAX_EXEMPLOS_DESCONHECIDOS = 20

# Todo registro do lote precisa ter EXATAMENTE as mesmas chaves: o insert
# multi-linha do SQLAlchemy compila um VALUES só para o lote inteiro, e um
# dict com chave a menos derruba a compilação. Por isso todo registro nasce
# deste molde e só depois é preenchido.
CAMPOS_LAVAGEM = (
    "id_externo", "origem", "importacao_id", "agendamento_id",
    "cliente_id", "veiculo_id", "tipo_carro",
    "funcionario_lavagem_id", "atendente_pagamento_id",
    "data", "chegou_em", "iniciou_lavagem_em", "terminou_lavagem_em", "saiu_em", "pago_em",
    "tempo_espera_antes_lavagem_min", "tempo_lavagem_total_min",
    "tempo_pos_lavagem_ate_retirada_min", "tempo_ate_pagamento_min",
    "t_teto_vidros_min", "t_capo_parabrisa_min", "t_laterais_portas_min",
    "t_traseira_min", "t_rodas_pneus_min", "t_interior_min",
    "servico", "preco_reais", "metodo_pagamento", "cnpj_receita",
    "nps_cliente", "nota_google",
    "shampoo_ml", "cera_ml", "pretinho_ml", "aromatizante_ml",
    "status", "observacoes", "dados_extras",
)


@dataclass
class ResultadoAba:
    """O que aconteceu com uma aba: contadores e avisos que vão para a tela."""

    aba: str
    lidas: int = 0
    inseridas: int = 0
    ignoradas: int = 0
    linhas_sem_data: int = 0
    datas_mm_dd: int = 0
    precos_invalidos: int = 0
    tipos_carro_desconhecidos: list[str] = field(default_factory=list)
    metodos_pagamento_desconhecidos: list[str] = field(default_factory=list)

    def anotar_desconhecido(self, lista: list[str], valor: str) -> None:
        if valor not in lista and len(lista) < MAX_EXEMPLOS_DESCONHECIDOS:
            lista.append(valor)

    def como_dict(self) -> dict:
        return {
            "aba": self.aba,
            "lidas": self.lidas,
            "inseridas": self.inseridas,
            "ignoradas": self.ignoradas,
            "avisos": {
                "linhas_sem_data": self.linhas_sem_data,
                "datas_mm_dd": self.datas_mm_dd,
                "precos_invalidos": self.precos_invalidos,
                "tipos_carro_desconhecidos": self.tipos_carro_desconhecidos,
                "metodos_pagamento_desconhecidos": self.metodos_pagamento_desconhecidos,
            },
        }


@dataclass
class ResultadoImportacao:
    """O resumo da carga inteira — vira o `detalhes` (JSONB) da importação."""

    modo: str  # "substituir" | "incremental"
    abas: list[ResultadoAba] = field(default_factory=list)
    abas_puladas: list[dict] = field(default_factory=list)
    lavagens_removidas: int = 0
    clientes_novos: int = 0
    funcionarios_novos: int = 0

    @property
    def lidas(self) -> int:
        return sum(aba.lidas for aba in self.abas)

    @property
    def inseridas(self) -> int:
        return sum(aba.inseridas for aba in self.abas)

    @property
    def ignoradas(self) -> int:
        return sum(aba.ignoradas for aba in self.abas)

    def como_detalhes(self) -> dict:
        return {
            "modo": self.modo,
            "lavagens_removidas": self.lavagens_removidas,
            "clientes_novos": self.clientes_novos,
            "funcionarios_novos": self.funcionarios_novos,
            "abas": [aba.como_dict() for aba in self.abas],
            "abas_puladas": self.abas_puladas,
        }


def registro_vazio() -> dict:
    molde = {campo: None for campo in CAMPOS_LAVAGEM}
    molde["servico"] = "Lavagem completa"  # NOT NULL no banco
    return molde


def canonizar(coluna) -> str:
    """Nome da coluna da planilha → nome canônico deste importador."""
    limpo = normalizar_texto(str(coluna)).replace(" ", "_")
    return APELIDOS.get(limpo, limpo)


def como_timestamp(valor, dia: date | None):
    """Aceita timestamp completo, ou só a hora (que é combinada com o dia)."""
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, time):
        return datetime.combine(dia, valor) if dia else None
    texto = str(valor).strip()
    try:
        return datetime.fromisoformat(texto)
    except ValueError:
        pass
    for formato in ("%H:%M:%S", "%H:%M"):
        try:
            hora = datetime.strptime(texto, formato).time()
            return datetime.combine(dia, hora) if dia else None
        except ValueError:
            continue
    return None


def nome_do_cliente(valor) -> str | None:
    """O nome como texto, ou None quando é o placeholder de nome ausente."""
    if not valor:
        return None
    nome = str(valor).strip()
    if not nome or normalizar_texto(nome) == NOME_PLACEHOLDER:
        return None
    return nome


def _inteiro(valor) -> int | None:
    """Célula numérica do Excel como int — openpyxl às vezes entrega float."""
    if valor is None:
        return None
    try:
        return int(float(valor))
    except (TypeError, ValueError):
        return None


@dataclass
class _Aba:
    """Uma aba bufferizada: cabeçalho resolvido e linhas em memória.

    Em memória porque o workbook read_only não rebobina, e o motor precisa de
    duas passadas (identidades, depois lavagens). São ~28 mil linhas no
    arquivo oficial — dezenas de MB, barato; se a base crescer uma ordem de
    grandeza, é aqui que se volta ao streaming com uma leitura por passada.
    """

    nome: str
    indice: dict[str, int]
    desconhecidas: list[str]
    linhas: list[tuple]

    def valor(self, linha: tuple, coluna: str):
        posicao = self.indice.get(coluna)
        if posicao is None or posicao >= len(linha):
            return None
        return linha[posicao]


def _ler_abas(origem, apenas_abas: list[str] | None) -> tuple[list[_Aba], list[dict]]:
    """Abre o workbook UMA vez e devolve (abas válidas, abas puladas)."""
    if isinstance(origem, (bytes, bytearray)):
        fonte = BytesIO(origem)
    else:
        fonte = Path(origem)
    wb = openpyxl.load_workbook(fonte, read_only=True, data_only=True)

    conhecidas = set(COLUNAS_DIRETAS) | set(COLUNAS_ESPECIAIS) | set(COLUNAS_IGNORADAS)
    abas: list[_Aba] = []
    puladas: list[dict] = []
    try:
        for nome in wb.sheetnames:
            if apenas_abas is not None and nome not in apenas_abas:
                continue
            ws = wb[nome]
            iterador = ws.iter_rows(values_only=True)
            try:
                cabecalho = [canonizar(c) if c is not None else "" for c in next(iterador)]
            except StopIteration:
                puladas.append({"aba": nome, "motivo": "aba vazia"})
                continue

            indice = {coluna: posicao for posicao, coluna in enumerate(cabecalho)}
            if not COLUNAS_OBRIGATORIAS <= set(indice):
                faltam = ", ".join(sorted(COLUNAS_OBRIGATORIAS - set(indice)))
                puladas.append(
                    {"aba": nome, "motivo": f"cabeçalho sem colunas obrigatórias ({faltam})"}
                )
                continue

            linhas = [
                linha for linha in iterador
                if linha is not None and any(v is not None for v in linha)
            ]
            desconhecidas = [c for c in indice if c and c not in conhecidas]
            abas.append(_Aba(nome=nome, indice=indice, desconhecidas=desconhecidas, linhas=linhas))
    finally:
        wb.close()
    return abas, puladas


def importar_planilha(
    origem,
    session,
    importacao_id: int,
    substituir: bool = False,
    apenas_abas: list[str] | None = None,
    tamanho_lote: int = LOTE_PADRAO,
) -> ResultadoImportacao:
    """Importa todas as abas de lavagens de um .xlsx (bytes ou caminho).

    Quem chama é dono da transação e do registro em `importacoes`: aqui não há
    commit em modo substituir até o fim, e o único commit intermediário do
    modo incremental é o de cada lote. Levanta exceção em erro — o chamador
    decide o que gravar no registro.
    """
    abas, puladas = _ler_abas(origem, apenas_abas)
    if not abas:
        motivos = "; ".join(f"{p['aba']}: {p['motivo']}" for p in puladas) or "arquivo sem abas"
        raise ValueError(f"nenhuma aba com as colunas esperadas (id_lavagem, data). {motivos}")

    resultado = ResultadoImportacao(
        modo="substituir" if substituir else "incremental",
        abas_puladas=puladas,
    )

    clientes = ClienteRepository(session)
    funcionarios = FuncionarioRepository(session)
    lavagens = LavagemRepository(session)

    # --- passada 1: quem é quem, cruzando todas as abas -------------------
    # Uma rodada só de upserts para o arquivo inteiro: cada rodada são 3
    # queries por entidade contra um banco a ~300 ms de distância.
    clientes_por_cpf: dict[str, str | None] = {}
    nomes_clientes: set[str] = set()
    nomes_funcionarios: set[str] = set()
    for aba in abas:
        tem_cpf = "cpf" in aba.indice
        for linha in aba.linhas:
            nome = nome_do_cliente(aba.valor(linha, "cliente"))
            if tem_cpf and aba.valor(linha, "cpf"):
                cpf = normalizar_cpf(str(aba.valor(linha, "cpf")))
                # O primeiro nome não-nulo do CPF ganha: a mesma pessoa aparece
                # em dezenas de linhas, e em algumas o nome pode faltar.
                if clientes_por_cpf.get(cpf) is None:
                    clientes_por_cpf[cpf] = nome
            elif nome:
                nomes_clientes.add(nome)
            for coluna in ("funcionario_lavagem", "atendente_pagamento"):
                valor = aba.valor(linha, coluna)
                if valor:
                    nomes_funcionarios.add(str(valor).strip())

    clientes_antes = session.query(func.count(ClienteModel.id)).scalar() or 0
    funcionarios_antes = funcionarios.session.query(
        func.count(funcionarios.model.id)
    ).scalar() or 0

    mapa_funcionarios = funcionarios.upsert_em_lote_por_nome(nomes_funcionarios)
    mapa_clientes_cpf = clientes.upsert_em_lote_por_cpf(clientes_por_cpf)
    mapa_clientes_nome = clientes.upsert_em_lote_por_nome(nomes_clientes)

    resultado.clientes_novos = (
        (session.query(func.count(ClienteModel.id)).scalar() or 0) - clientes_antes
    )
    resultado.funcionarios_novos = (
        (funcionarios.session.query(func.count(funcionarios.model.id)).scalar() or 0)
        - funcionarios_antes
    )

    # --- substituir: sai a carga antiga, na mesma transação ---------------
    if substituir:
        resultado.lavagens_removidas = lavagens.apagar_importadas()

    # --- passada 2: as lavagens, aba a aba --------------------------------
    for aba in abas:
        parcial = ResultadoAba(aba=aba.nome)
        resultado.abas.append(parcial)
        tem_cpf = "cpf" in aba.indice
        lote: list[dict] = []

        def descarregar() -> None:
            parcial.inseridas += lavagens.inserir_lote_importado(lote)
            lote.clear()
            if not substituir:
                session.commit()

        for linha in aba.linhas:
            parcial.lidas += 1
            registro, extras = registro_vazio(), {}

            dia, foi_mm_dd = padronizar_data(
                aba.valor(linha, "data"),
                ano=_inteiro(aba.valor(linha, "ano")),
                mes=_inteiro(aba.valor(linha, "mes")),
            )
            if dia is None:
                parcial.linhas_sem_data += 1
                continue  # sem data não há como colocar num período
            if foi_mm_dd:
                parcial.datas_mm_dd += 1
            registro["data"] = dia

            for coluna in COLUNAS_DIRETAS:
                if coluna in aba.indice:
                    registro[coluna] = aba.valor(linha, coluna)

            tipo, conhecido = padronizar_categoria(
                aba.valor(linha, "tipo_carro"), TIPOS_CARRO_CANONICOS
            )
            registro["tipo_carro"] = tipo
            if not conhecido:
                parcial.anotar_desconhecido(
                    parcial.tipos_carro_desconhecidos, str(aba.valor(linha, "tipo_carro"))
                )

            metodo, conhecido = padronizar_categoria(
                aba.valor(linha, "metodo_pagamento"), METODOS_PAGAMENTO_CANONICOS
            )
            registro["metodo_pagamento"] = metodo
            if not conhecido:
                parcial.anotar_desconhecido(
                    parcial.metodos_pagamento_desconhecidos,
                    str(aba.valor(linha, "metodo_pagamento")),
                )

            preco_bruto = aba.valor(linha, "preco_reais")
            registro["preco_reais"] = padronizar_preco(preco_bruto)
            if preco_bruto is not None and registro["preco_reais"] is None:
                parcial.precos_invalidos += 1

            for coluna in ("chegou_em", "iniciou_lavagem_em", "terminou_lavagem_em",
                           "saiu_em", "pago_em"):
                if coluna in aba.indice:
                    registro[coluna] = como_timestamp(aba.valor(linha, coluna), dia)

            registro["id_externo"] = _inteiro(aba.valor(linha, "id_lavagem"))
            if aba.valor(linha, "servico"):
                registro["servico"] = str(aba.valor(linha, "servico")).strip()

            # cliente: o CPF manda; sem ele, casa pelo nome
            if tem_cpf and aba.valor(linha, "cpf"):
                registro["cliente_id"] = mapa_clientes_cpf.get(
                    normalizar_cpf(str(aba.valor(linha, "cpf")))
                )
            else:
                nome_cliente = nome_do_cliente(aba.valor(linha, "cliente"))
                if nome_cliente:
                    registro["cliente_id"] = mapa_clientes_nome.get(
                        normalizar_texto(nome_cliente)
                    )

            for coluna_origem, destino in (
                ("funcionario_lavagem", "funcionario_lavagem_id"),
                ("atendente_pagamento", "atendente_pagamento_id"),
            ):
                valor = aba.valor(linha, coluna_origem)
                if valor:
                    registro[destino] = mapa_funcionarios.get(
                        normalizar_texto(str(valor).strip())
                    )

            # telefone e placa: hoje ausentes, tratados para quando chegarem
            if aba.valor(linha, "telefone"):
                extras["telefone"] = normalizar_telefone(str(aba.valor(linha, "telefone")))
            if aba.valor(linha, "placa"):
                extras["placa"] = normalizar_placa(str(aba.valor(linha, "placa")))
            if aba.valor(linha, "modelo_carro"):
                extras["modelo_carro"] = str(aba.valor(linha, "modelo_carro")).strip()

            for coluna in aba.desconhecidas:
                valor = aba.valor(linha, coluna)
                if valor is not None:
                    extras[coluna] = valor if isinstance(valor, (int, float, str)) else str(valor)

            # se vieram horários mas não vieram durações, calcula — a mesma
            # regra que a plataforma aplica no check-out
            for inicio, fim, duracao in (
                ("chegou_em", "iniciou_lavagem_em", "tempo_espera_antes_lavagem_min"),
                ("iniciou_lavagem_em", "terminou_lavagem_em", "tempo_lavagem_total_min"),
                ("terminou_lavagem_em", "saiu_em", "tempo_pos_lavagem_ate_retirada_min"),
                ("terminou_lavagem_em", "pago_em", "tempo_ate_pagamento_min"),
            ):
                if registro.get(duracao) is None:
                    calculado = minutos_entre(registro.get(inicio), registro.get(fim))
                    if calculado is not None:
                        registro[duracao] = calculado

            registro["origem"] = "importacao"
            registro["status"] = "concluida"
            registro["importacao_id"] = importacao_id
            registro["dados_extras"] = extras

            lote.append(registro)
            if len(lote) >= tamanho_lote:
                descarregar()

        # o lote nunca cruza a fronteira da aba: é o que mantém o `inseridas`
        # de cada aba exato mesmo com o DO NOTHING engolindo repetidas
        descarregar()
        parcial.ignoradas = parcial.lidas - parcial.linhas_sem_data - parcial.inseridas

    # --- veículos, quando a base trouxer placa ----------------------------
    criar_veiculos_dos_extras(session, importacao_id)

    session.commit()
    return resultado


def criar_veiculos_dos_extras(session, importacao_id: int) -> int:
    """Cria veículos para as lavagens cuja planilha trouxe placa.

    Roda depois da carga porque só aí os clientes já têm id. Com a base de
    hoje não faz nada — ela não tem placa. Existe para o dia em que tiver.
    Não comita: faz parte da transação da carga.
    """
    lavagens = (
        session.query(LavagemModel)
        .filter(
            LavagemModel.importacao_id == importacao_id,
            LavagemModel.dados_extras.has_key("placa"),  # noqa: W601
        )
        .all()
    )
    if not lavagens:
        return 0

    criados = 0
    cache: dict[str, int] = {}
    for lavagem in lavagens:
        placa = lavagem.dados_extras.get("placa")
        if not placa:
            continue
        if placa not in cache:
            veiculo = session.query(VeiculoModel).filter(VeiculoModel.placa == placa).first()
            if veiculo is None:
                veiculo = VeiculoModel(
                    placa=placa,
                    cliente_id=lavagem.cliente_id,
                    tipo=lavagem.tipo_carro,
                    modelo=lavagem.dados_extras.get("modelo_carro"),
                )
                session.add(veiculo)
                session.flush()
                criados += 1
            cache[placa] = veiculo.id
        lavagem.veiculo_id = cache[placa]

    # telefone que veio na planilha completa o cadastro do cliente
    for lavagem in lavagens:
        telefone = lavagem.dados_extras.get("telefone")
        if telefone and lavagem.cliente_id:
            cliente = session.get(ClienteModel, lavagem.cliente_id)
            if cliente is not None and not cliente.telefone:
                cliente.telefone = telefone

    return criados


def processar_upload(importacao_id: int, conteudo: bytes, substituir: bool) -> None:
    """A carga como o BackgroundTasks a executa, depois da resposta 202.

    Sessão PRÓPRIA, não a do request — aquela fecha junto com a resposta. Os
    bytes já chegaram lidos pelo endpoint pelo mesmo motivo: o UploadFile
    também morre com o request. O registro em `importacoes` é atualizado fora
    da transação de dados, então o polling vê 'processando' até o commit da
    carga e o resultado logo depois.
    """
    session = SessionLocal()
    try:
        resultado = importar_planilha(
            conteudo, session, importacao_id, substituir=substituir
        )
        importacao = session.get(ImportacaoModel, importacao_id)
        importacao.status = "concluida"
        importacao.linhas_lidas = resultado.lidas
        importacao.linhas_inseridas = resultado.inseridas
        importacao.linhas_ignoradas = resultado.ignoradas
        importacao.detalhes = resultado.como_detalhes()
        importacao.concluida_em = datetime.now(timezone.utc)
        session.commit()
    except Exception as erro:  # noqa: BLE001 — o job não tem quem o observe além do registro
        session.rollback()
        try:
            importacao = session.get(ImportacaoModel, importacao_id)
            importacao.status = "erro"
            importacao.erro = str(erro)[:2000]
            importacao.concluida_em = datetime.now(timezone.utc)
            session.commit()
        except Exception:
            session.rollback()
    finally:
        session.close()

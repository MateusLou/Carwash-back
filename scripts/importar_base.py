"""Importa uma planilha de lavagens para o schema `operacao`.

    python scripts/importar_base.py --arquivo "../base tratada.xlsx"

Feito para sobreviver à base que ainda vai chegar. Três mecanismos:

1. **Mapa de colunas explícito** — o que a base de hoje tem.
2. **Apelidos** — nomes prováveis das colunas que hoje faltam (telefone, placa,
   horário de chegada e de saída). Se vierem com um desses nomes, entram na
   coluna certa sozinhas.
3. **`dados_extras`** — qualquer coluna desconhecida vai para o jsonb, e o script
   diz no fim quais foram. Nenhuma importação falha por causa de coluna nova;
   depois se decide quais merecem coluna própria (via migration do Alembic).

Reimportar o mesmo arquivo não duplica nada: `id_externo` (o `id_lavagem` da
planilha) tem índice único e o insert usa ON CONFLICT DO NOTHING.

O cliente é deduplicado por **CPF** quando a planilha traz a coluna, e só cai
para o nome quando não traz. Dedup por nome perde gente: na base imputada são 795
CPFs contra 789 nomes, porque sete pessoas ficaram com o nome "Não informado".

Para recarregar do zero, `--recomecar` esvazia as tabelas do schema `operacao`
antes de importar. Sem ele o script nunca apaga nada.
"""

import argparse
import os
import sys
from datetime import date, datetime, time
from pathlib import Path

import openpyxl
from sqlalchemy import select, text
from sqlalchemy.dialects.postgresql import insert

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from database.database import SessionLocal  # noqa: E402
from models.cliente_model import ClienteModel  # noqa: E402
from models.funcionario_model import FuncionarioModel  # noqa: E402
from models.importacao_model import ImportacaoModel  # noqa: E402
from models.lavagem_model import LavagemModel  # noqa: E402
from models.veiculo_model import VeiculoModel  # noqa: E402
from entities.lavagem import minutos_entre  # noqa: E402
from repositories.veiculo_repository import normalizar_placa  # noqa: E402
from utils.normalizar_texto import (  # noqa: E402
    normalizar_cpf,
    normalizar_telefone,
    normalizar_texto,
)


# Colunas que viram coluna de mesmo nome na tabela de lavagens.
COLUNAS_DIRETAS = (
    "tipo_carro",
    "tempo_espera_antes_lavagem_min",
    "tempo_lavagem_total_min",
    "tempo_pos_lavagem_ate_retirada_min",
    "tempo_ate_pagamento_min",
    "t_teto_vidros_min",
    "t_capo_parabrisa_min",
    "t_laterais_portas_min",
    "t_traseira_min",
    "t_rodas_pneus_min",
    "t_interior_min",
    "preco_reais",
    "metodo_pagamento",
    "cnpj_receita",
    "nps_cliente",
    "nota_google",
    "shampoo_ml",
    "cera_ml",
    "pretinho_ml",
    "aromatizante_ml",
    "observacoes",
)

# Colunas com tratamento próprio (viram FK, ou têm nome diferente no banco).
COLUNAS_ESPECIAIS = {
    "id_lavagem": "id_externo",
    "data": "data",
    "cpf": "cpf",
    "cliente": "cliente",
    "telefone": "telefone",
    "placa": "placa",
    "modelo_carro": "modelo",
    "funcionario_lavagem": "funcionario_lavagem",
    "atendente_pagamento": "atendente_pagamento",
    "servico": "servico",
    "chegou_em": "chegou_em",
    "iniciou_lavagem_em": "iniciou_lavagem_em",
    "terminou_lavagem_em": "terminou_lavagem_em",
    "saiu_em": "saiu_em",
    "pago_em": "pago_em",
}

# Apelidos: como as colunas que hoje faltam provavelmente virão escritas.
APELIDOS = {
    "telefone_cliente": "telefone",
    "celular": "telefone",
    "whatsapp": "telefone",
    "placa_veiculo": "placa",
    "modelo": "modelo_carro",
    "horario_chegada": "chegou_em",
    "hora_chegada": "chegou_em",
    "chegada": "chegou_em",
    "horario_entrada": "chegou_em",
    "horario_inicio_lavagem": "iniciou_lavagem_em",
    "horario_fim_lavagem": "terminou_lavagem_em",
    "horario_saida": "saiu_em",
    "hora_saida": "saiu_em",
    "saida": "saiu_em",
    "horario_pagamento": "pago_em",
    "nome_cliente": "cliente",
    "lavador": "funcionario_lavagem",
    "atendente": "atendente_pagamento",
    "valor": "preco_reais",
    "valor_reais": "preco_reais",
    "preco": "preco_reais",
    "nps": "nps_cliente",
}

# Derivadas de `data` — não viram coluna, para não haver duas versões do mesmo
# fato no banco (o dashboard tira ano/mês/dia da semana com date_trunc).
COLUNAS_IGNORADAS = ("ano", "mes", "mês", "dia_semana")

# O que a imputação escreveu no lugar do nome que faltava. Guardar isso como
# nome real criaria 7 cadastros homônimos falsos — é um nulo, e entra como nulo.
NOME_PLACEHOLDER = "nao informado"

# Tabelas que este script possui. O `--recomecar` esvazia exatamente estas, e
# nenhuma do `public` — lá moram os agendamentos e conversas do bot e o login.
TABELAS_DA_PLATAFORMA = (
    "operacao.lavagens",
    "operacao.veiculos",
    "operacao.clientes",
    "operacao.funcionarios",
    "operacao.importacoes",
)

LOTE_PADRAO = 5000

# Todo registro do lote precisa ter EXATAMENTE as mesmas chaves: o insert
# multi-linha do SQLAlchemy compila um VALUES só para o lote inteiro, e um dict
# com chave a menos derruba a compilação. Por isso todo registro nasce deste
# molde e só depois é preenchido.
CAMPOS_LAVAGEM = (
    "id_externo", "origem", "importacao_id", "agendamento_id",
    "cliente_id", "veiculo_id", "tipo_carro",
    "funcionario_lavagem_id", "atendente_pagamento_id",
    "data", "chegou_em", "iniciou_lavagem_em", "terminou_lavagem_em", "saiu_em", "pago_em",
    "tempo_espera_antes_lavagem_min", "tempo_lavagem_total_min",
    "tempo_pos_lavagem_ate_retirada_min", "tempo_ate_pagamento_min",
    "t_teto_vidros_min", "t_capo_parabrisa_min", "t_laterais_portas_min",
    "t_traseira_min", "t_rodas_pneus_min", "t_interior_min",
    "servico", "preco_reais", "metodo_pagamento", "cnpj_receita",
    "nps_cliente", "nota_google",
    "shampoo_ml", "cera_ml", "pretinho_ml", "aromatizante_ml",
    "status", "observacoes", "dados_extras",
)


def registro_vazio() -> dict:
    molde = {campo: None for campo in CAMPOS_LAVAGEM}
    molde["servico"] = "Lavagem completa"  # NOT NULL no banco
    return molde


def canonizar(coluna: str) -> str:
    """Nome da coluna da planilha → nome canônico deste importador."""
    limpo = normalizar_texto(str(coluna)).replace(" ", "_")
    return APELIDOS.get(limpo, limpo)


def como_data(valor) -> date | None:
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return date.fromisoformat(str(valor).strip()[:10])


def como_timestamp(valor, dia: date | None):
    """Aceita timestamp completo, ou só a hora (que é combinada com o dia)."""
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, time):
        return datetime.combine(dia, valor) if dia else None
    texto = str(valor).strip()
    try:
        return datetime.fromisoformat(texto)
    except ValueError:
        pass
    for formato in ("%H:%M:%S", "%H:%M"):
        try:
            hora = datetime.strptime(texto, formato).time()
            return datetime.combine(dia, hora) if dia else None
        except ValueError:
            continue
    return None


def ler_cabecalho(caminho: Path, aba: str) -> tuple[list[str], list[str]]:
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[aba]
    bruto = [str(c) if c is not None else "" for c in next(ws.iter_rows(values_only=True))]
    wb.close()
    return bruto, [canonizar(c) for c in bruto]


def linhas_da_planilha(caminho: Path, aba: str):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[aba]
    iterador = ws.iter_rows(values_only=True)
    next(iterador)  # cabeçalho
    for linha in iterador:
        if linha is None or all(v is None for v in linha):
            continue
        yield linha
    wb.close()


def upsert_por_nome(session, model, nomes: set[str]) -> dict[str, int]:
    """Garante que cada nome exista na tabela e devolve nome_normalizado → id.

    Em três queries, não uma por linha: com 3 mil clientes na base, o
    get_or_create linha a linha levaria minutos contra o Supabase.
    """
    normalizados = {normalizar_texto(n): n for n in nomes if n and str(n).strip()}
    if not normalizados:
        return {}

    existentes = dict(
        session.execute(
            select(model.nome_normalizado, model.id).where(
                model.nome_normalizado.in_(list(normalizados))
            )
        ).all()
    )

    faltando = [
        {"nome": original, "nome_normalizado": chave}
        for chave, original in normalizados.items()
        if chave not in existentes
    ]
    if faltando:
        session.execute(insert(model).values(faltando).on_conflict_do_nothing())
        session.commit()
        existentes = dict(
            session.execute(
                select(model.nome_normalizado, model.id).where(
                    model.nome_normalizado.in_(list(normalizados))
                )
            ).all()
        )
    return existentes


def nome_do_cliente(valor) -> str | None:
    """O nome como texto, ou None quando é o placeholder da imputação.

    "Não informado" é o que a imputação escreveu onde o nome faltava. É um nulo
    escrito por extenso, e entra no banco como nulo.
    """
    if not valor:
        return None
    nome = str(valor).strip()
    if not nome or normalizar_texto(nome) == NOME_PLACEHOLDER:
        return None
    return nome


def upsert_clientes_por_cpf(session, pares: dict[str, str | None]) -> dict[str, int]:
    """Garante que cada CPF exista em clientes e devolve cpf → id.

    `pares` é cpf normalizado → nome (ou None). Mesmas três queries do
    `upsert_por_nome`, pela mesma razão: um get_or_create por linha levaria
    minutos contra o Supabase.

    O nome entra junto, mas quem manda é o CPF — dois homônimos de CPF diferente
    são duas linhas, e é para isso que o índice do nome deixou de ser único.
    """
    if not pares:
        return {}

    existentes = dict(
        session.execute(
            select(ClienteModel.cpf, ClienteModel.id).where(
                ClienteModel.cpf.in_(list(pares))
            )
        ).all()
    )

    faltando = [
        {
            "cpf": cpf,
            "nome": nome,
            "nome_normalizado": normalizar_texto(nome) if nome else None,
        }
        for cpf, nome in pares.items()
        if cpf not in existentes
    ]
    if faltando:
        session.execute(insert(ClienteModel).values(faltando).on_conflict_do_nothing())
        session.commit()
        existentes = dict(
            session.execute(
                select(ClienteModel.cpf, ClienteModel.id).where(
                    ClienteModel.cpf.in_(list(pares))
                )
            ).all()
        )
    return existentes


def recomecar(session) -> None:
    """Esvazia as tabelas da plataforma para uma carga do zero.

    Destrutivo e irreversível, por isso atrás de confirmação digitada. O
    `public` fica intacto: agendamentos e conversas do bot, e o login, não são
    deste script. RESTART IDENTITY para os ids voltarem a começar do 1.
    """
    print("\n⚠️  --recomecar vai APAGAR, no schema operacao:")
    for tabela in TABELAS_DA_PLATAFORMA:
        total = session.execute(text(f"select count(*) from {tabela}")).scalar()
        print(f"   {tabela:<24} {total:>10,} linhas")
    print("   (o schema public — agendamentos, conversas, login — não é tocado)")

    if input('\ndigite "RECOMECAR" para confirmar: ').strip() != "RECOMECAR":
        print("cancelado.")
        raise SystemExit(1)

    session.execute(
        text(f"truncate {', '.join(TABELAS_DA_PLATAFORMA)} restart identity cascade")
    )
    session.commit()
    print("tabelas esvaziadas.\n")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--arquivo", required=True, help="caminho do .xlsx")
    parser.add_argument("--aba", default="lavagens")
    parser.add_argument("--lote", type=int, default=LOTE_PADRAO)
    parser.add_argument("--limite", type=int, default=None, help="importa só as N primeiras linhas")
    parser.add_argument(
        "--recomecar",
        action="store_true",
        help="APAGA as tabelas do schema operacao antes de importar",
    )
    args = parser.parse_args()

    caminho = Path(args.arquivo).expanduser().resolve()
    if not caminho.exists():
        print(f"arquivo não encontrado: {caminho}")
        return 1

    cabecalho_bruto, cabecalho = ler_cabecalho(caminho, args.aba)
    conhecidas = set(COLUNAS_DIRETAS) | set(COLUNAS_ESPECIAIS) | set(COLUNAS_IGNORADAS)
    desconhecidas = [c for c in cabecalho if c not in conhecidas]
    indice = {nome: i for i, nome in enumerate(cabecalho)}

    print(f"arquivo ....... {caminho.name}")
    print(f"colunas ....... {len(cabecalho)}")
    if desconhecidas:
        print(f"não mapeadas .. {', '.join(desconhecidas)}  → vão para dados_extras")
    for esperada in ("telefone", "placa", "chegou_em", "saiu_em"):
        if esperada not in indice:
            print(f"ausente ....... {esperada} (fica nulo; o check-in preenche daqui pra frente)")

    session = SessionLocal()
    if args.recomecar:
        recomecar(session)

    importacao = ImportacaoModel(arquivo=caminho.name, aba=args.aba)
    session.add(importacao)
    session.commit()
    session.refresh(importacao)

    try:
        # --- passada 1: quem é quem ------------------------------------
        tem_cpf = "cpf" in indice
        clientes_por_cpf: dict[str, str | None] = {}
        nomes_clientes: set[str] = set()
        nomes_funcionarios: set[str] = set()
        for linha in linhas_da_planilha(caminho, args.aba):
            nome = nome_do_cliente(linha[indice["cliente"]] if "cliente" in indice else None)
            if tem_cpf and linha[indice["cpf"]]:
                cpf = normalizar_cpf(str(linha[indice["cpf"]]))
                # O primeiro nome não-nulo do CPF ganha: a mesma pessoa aparece
                # em dezenas de linhas, e em algumas o nome falta.
                if clientes_por_cpf.get(cpf) is None:
                    clientes_por_cpf[cpf] = nome
            elif nome:
                nomes_clientes.add(nome)
            for coluna in ("funcionario_lavagem", "atendente_pagamento"):
                if coluna in indice and linha[indice[coluna]]:
                    nomes_funcionarios.add(str(linha[indice[coluna]]).strip())

        mapa_funcionarios = upsert_por_nome(session, FuncionarioModel, nomes_funcionarios)
        mapa_clientes_cpf = upsert_clientes_por_cpf(session, clientes_por_cpf)
        mapa_clientes_nome = upsert_por_nome(session, ClienteModel, nomes_clientes)
        chave = "CPF" if tem_cpf else "nome (a planilha não tem CPF)"
        print(f"funcionários .. {len(mapa_funcionarios)}")
        print(f"clientes ...... {len(mapa_clientes_cpf) + len(mapa_clientes_nome)} por {chave}")

        # --- passada 2: as lavagens ------------------------------------
        lote: list[dict] = []
        lidas = inseridas = 0

        def descarregar() -> int:
            if not lote:
                return 0
            comando = (
                insert(LavagemModel)
                .values(lote)
                .on_conflict_do_nothing(
                    index_elements=["id_externo"],
                    index_where=text("id_externo is not null"),
                )
            )
            resultado = session.execute(comando)
            session.commit()
            lote.clear()
            return resultado.rowcount or 0

        for linha in linhas_da_planilha(caminho, args.aba):
            if args.limite is not None and lidas >= args.limite:
                break
            lidas += 1
            registro, extras = registro_vazio(), {}

            dia = como_data(linha[indice["data"]]) if "data" in indice else None
            if dia is None:
                continue  # sem data não há como colocar num período
            registro["data"] = dia

            for coluna in COLUNAS_DIRETAS:
                if coluna in indice:
                    registro[coluna] = linha[indice[coluna]]

            for coluna in ("chegou_em", "iniciou_lavagem_em", "terminou_lavagem_em",
                           "saiu_em", "pago_em"):
                if coluna in indice:
                    registro[coluna] = como_timestamp(linha[indice[coluna]], dia)

            if "id_lavagem" in indice:
                registro["id_externo"] = linha[indice["id_lavagem"]]
            if "servico" in indice and linha[indice["servico"]]:
                registro["servico"] = linha[indice["servico"]]

            # cliente: o CPF manda; sem ele, casa pelo nome
            if tem_cpf and linha[indice["cpf"]]:
                registro["cliente_id"] = mapa_clientes_cpf.get(
                    normalizar_cpf(str(linha[indice["cpf"]]))
                )
            else:
                nome_cliente = nome_do_cliente(
                    linha[indice["cliente"]] if "cliente" in indice else None
                )
                if nome_cliente:
                    registro["cliente_id"] = mapa_clientes_nome.get(
                        normalizar_texto(nome_cliente)
                    )

            for origem, destino in (("funcionario_lavagem", "funcionario_lavagem_id"),
                                    ("atendente_pagamento", "atendente_pagamento_id")):
                if origem in indice and linha[indice[origem]]:
                    registro[destino] = mapa_funcionarios.get(
                        normalizar_texto(str(linha[indice[origem]]).strip())
                    )

            # telefone e placa: hoje ausentes, tratados para quando chegarem
            if "telefone" in indice and linha[indice["telefone"]]:
                extras["telefone"] = normalizar_telefone(str(linha[indice["telefone"]]))
            if "placa" in indice and linha[indice["placa"]]:
                extras["placa"] = normalizar_placa(str(linha[indice["placa"]]))
            if "modelo_carro" in indice and linha[indice["modelo_carro"]]:
                extras["modelo_carro"] = str(linha[indice["modelo_carro"]]).strip()

            for coluna in desconhecidas:
                valor = linha[indice[coluna]]
                if valor is not None:
                    extras[coluna] = valor if isinstance(valor, (int, float, str)) else str(valor)

            # se vieram horários mas não vieram durações, calcula — a mesma
            # regra que a plataforma aplica no check-out
            for inicio, fim, duracao in (
                ("chegou_em", "iniciou_lavagem_em", "tempo_espera_antes_lavagem_min"),
                ("iniciou_lavagem_em", "terminou_lavagem_em", "tempo_lavagem_total_min"),
                ("terminou_lavagem_em", "saiu_em", "tempo_pos_lavagem_ate_retirada_min"),
                ("terminou_lavagem_em", "pago_em", "tempo_ate_pagamento_min"),
            ):
                if registro.get(duracao) is None:
                    calculado = minutos_entre(registro.get(inicio), registro.get(fim))
                    if calculado is not None:
                        registro[duracao] = calculado

            registro["origem"] = "importacao"
            registro["status"] = "concluida"
            registro["importacao_id"] = importacao.id
            registro["dados_extras"] = extras

            lote.append(registro)
            if len(lote) >= args.lote:
                inseridas += descarregar()
                print(f"  {lidas:>8,} lidas | {inseridas:>8,} inseridas", end="\r", flush=True)

        inseridas += descarregar()

        # --- veículos, quando a base trouxer placa ---------------------
        veiculos = criar_veiculos_dos_extras(session, importacao.id)

        importacao.linhas_lidas = lidas
        importacao.linhas_inseridas = inseridas
        importacao.linhas_ignoradas = lidas - inseridas
        importacao.concluida_em = datetime.now()
        session.commit()

        print(" " * 60, end="\r")
        print(f"lidas ......... {lidas:,}")
        print(f"inseridas ..... {inseridas:,}")
        print(f"ignoradas ..... {lidas - inseridas:,} (já estavam no banco, por id_externo)")
        if veiculos:
            print(f"veículos ...... {veiculos:,} criados a partir da placa")
        print(f"importação #{importacao.id} concluída")
        return 0

    except Exception as erro:  # noqa: BLE001
        session.rollback()
        importacao.erro = str(erro)[:2000]
        importacao.concluida_em = datetime.now()
        session.commit()
        print(f"\nfalhou: {erro}")
        raise
    finally:
        session.close()


def criar_veiculos_dos_extras(session, importacao_id: int) -> int:
    """Cria veículos para as lavagens cuja planilha trouxe placa.

    Roda depois da carga porque só aí os clientes já têm id. Com a base de hoje
    não faz nada — ela não tem placa. Existe para o dia em que tiver.
    """
    lavagens = (
        session.query(LavagemModel)
        .filter(
            LavagemModel.importacao_id == importacao_id,
            LavagemModel.dados_extras.has_key("placa"),  # noqa: W601
        )
        .all()
    )
    if not lavagens:
        return 0

    criados = 0
    cache: dict[str, int] = {}
    for lavagem in lavagens:
        placa = lavagem.dados_extras.get("placa")
        if not placa:
            continue
        if placa not in cache:
            veiculo = session.query(VeiculoModel).filter(VeiculoModel.placa == placa).first()
            if veiculo is None:
                veiculo = VeiculoModel(
                    placa=placa,
                    cliente_id=lavagem.cliente_id,
                    tipo=lavagem.tipo_carro,
                    modelo=lavagem.dados_extras.get("modelo_carro"),
                )
                session.add(veiculo)
                session.flush()
                criados += 1
            cache[placa] = veiculo.id
        lavagem.veiculo_id = cache[placa]

    # telefone que veio na planilha completa o cadastro do cliente
    for lavagem in lavagens:
        telefone = lavagem.dados_extras.get("telefone")
        if telefone and lavagem.cliente_id:
            cliente = session.query(ClienteModel).get(lavagem.cliente_id)
            if cliente is not None and not cliente.telefone:
                cliente.telefone = telefone

    session.commit()
    return criados


if __name__ == "__main__":
    os.environ.setdefault("PYTHONUNBUFFERED", "1")
    raise SystemExit(main())
